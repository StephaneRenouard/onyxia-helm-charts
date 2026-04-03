#!/usr/bin/env python3
import csv
import html
import io
import os
from pathlib import Path

from flask import Flask, redirect, request, Response
from openpyxl import Workbook, load_workbook
import xlrd
import xlwt

ROOT = Path(os.environ.get("PREMYOM_S3_MOUNT_ROOT", "/mnt/s3")).resolve()
MAX_FILE_SIZE = int(os.environ.get("TABULAR_EDITOR_MAX_BYTES", str(20 * 1024 * 1024)))
ALLOWED_EXTENSIONS = {".csv", ".xls", ".xlsx"}

app = Flask(__name__)


def escape(s: str) -> str:
    return html.escape(s, quote=True)


def safe_target(raw: str | None) -> Path | None:
    if not raw:
        return None
    candidate = (ROOT / raw).resolve()
    if ROOT != candidate and ROOT not in candidate.parents:
        return None
    return candidate


def split_ext(path: Path) -> str:
    return path.suffix.lower()


def rel_path(path: Path) -> str:
    return str(path.relative_to(ROOT))


def parse_tabular_text(content: str) -> list[list[str]]:
    stream = io.StringIO(content)
    return [list(row) for row in csv.reader(stream)]


def rows_to_csv(rows: list[list[str]]) -> str:
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerows(rows)
    return out.getvalue()


def read_rows(path: Path) -> list[list[str]]:
    ext = split_ext(path)
    if ext == ".csv":
        with path.open("r", encoding="utf-8", newline="") as f:
            return [list(row) for row in csv.reader(f)]
    if ext == ".xlsx":
        wb = load_workbook(filename=path, read_only=True, data_only=False)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if value is None else str(value)) for value in row])
        return rows
    if ext == ".xls":
        workbook = xlrd.open_workbook(path.as_posix())
        sheet = workbook.sheet_by_index(0)
        rows = []
        for row_index in range(sheet.nrows):
            rows.append([str(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)])
        return rows
    raise ValueError("unsupported file type")


def write_rows(path: Path, rows: list[list[str]]) -> None:
    ext = split_ext(path)
    if ext == ".csv":
        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, lineterminator="\n")
            writer.writerows(rows)
        return
    if ext == ".xlsx":
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(path.as_posix())
        return
    if ext == ".xls":
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        for i, row in enumerate(rows):
            for j, value in enumerate(row):
                ws.write(i, j, value)
        wb.save(path.as_posix())
        return
    raise ValueError("unsupported file type")


def directory_listing(base: Path) -> str:
    dirs: list[str] = []
    files: list[str] = []
    for entry in sorted(base.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower())):
        rel = rel_path(entry)
        if entry.is_dir():
            dirs.append(f'<li>📁 <a href="/tabular/?dir={escape(rel)}">{escape(entry.name)}</a></li>')
        elif split_ext(entry) in ALLOWED_EXTENSIONS:
            files.append(f'<li>📄 <a href="/tabular/?file={escape(rel)}">{escape(entry.name)}</a></li>')
    parent = ""
    if base != ROOT:
        parent_rel = rel_path(base.parent)
        parent = f'<li>⬆️ <a href="/tabular/?dir={escape(parent_rel)}">..</a></li>'
    return parent + "".join(dirs + files)


def page(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{escape(title)}</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, Segoe UI, Roboto, sans-serif; margin: 0; background: #111; color: #eee; }}
    .wrap {{ max-width: 1400px; margin: 0 auto; padding: 16px; }}
    a {{ color: #7dc4ff; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .box {{ background: #1b1b1b; border: 1px solid #333; border-radius: 8px; padding: 12px; margin-bottom: 12px; }}
    textarea {{ width: 100%; min-height: 62vh; background: #0f0f0f; color: #fff; border: 1px solid #333; border-radius: 6px; font-family: ui-monospace, SFMono-Regular, Menlo, monospace; font-size: 13px; }}
    button {{ background: #1677ff; color: #fff; border: none; border-radius: 6px; padding: 8px 12px; cursor: pointer; }}
    button[disabled] {{ background: #555; cursor: not-allowed; }}
    code {{ color: #ffd479; }}
    ul {{ margin: 0; padding-left: 20px; }}
  </style>
</head>
<body>
  <div class="wrap">
    {body}
  </div>
</body>
</html>"""


@app.get("/")
def root_redirect():
    return redirect("/tabular/", code=302)


@app.get("/tabular/")
def tabular_home():
    raw_dir = request.args.get("dir", "")
    raw_file = request.args.get("file")
    target_dir = safe_target(raw_dir) if raw_dir else ROOT
    if target_dir is None or not target_dir.exists() or not target_dir.is_dir():
        return Response(page("Erreur", '<div class="box">Répertoire invalide.</div>'), status=400, mimetype="text/html")

    if raw_file:
        file_path = safe_target(raw_file)
        if file_path is None or not file_path.exists() or not file_path.is_file():
            return Response(page("Erreur", '<div class="box">Fichier invalide.</div>'), status=400, mimetype="text/html")
        if split_ext(file_path) not in ALLOWED_EXTENSIONS:
            return Response(page("Erreur", '<div class="box">Type de fichier non supporté (csv/xls/xlsx).</div>'), status=400, mimetype="text/html")
        if file_path.stat().st_size > MAX_FILE_SIZE:
            return Response(page("Erreur", f'<div class="box">Fichier trop volumineux (&gt; {MAX_FILE_SIZE} octets).</div>'), status=400, mimetype="text/html")
        rows = read_rows(file_path)
        content = rows_to_csv(rows)
        writable = os.access(file_path, os.W_OK)
        disabled = "" if writable else "disabled"
        write_notice = "" if writable else "<div class=\"box\">Mode lecture seule (pas de droits RW sur ce mount).</div>"
        body = f"""
<div class="box">
  <a href="/tabular/?dir={escape(rel_path(file_path.parent))}">⬅ retour</a>
  <h2>Édition tabulaire: <code>{escape(rel_path(file_path))}</code></h2>
  <div>Format source: <code>{escape(split_ext(file_path))}</code> (affiché en CSV).</div>
</div>
{write_notice}
<form class="box" method="post" action="/tabular/save?file={escape(rel_path(file_path))}">
  <textarea name="content">{escape(content)}</textarea>
  <div style="margin-top:10px;">
    <button type="submit" {disabled}>Enregistrer</button>
    <span style="margin-left: 10px;">Le contenu est sauvegardé dans le format d'origine.</span>
  </div>
</form>
"""
        return Response(page("Éditeur tabulaire", body), mimetype="text/html")

    listing = directory_listing(target_dir)
    body = f"""
<div class="box">
  <h2>Éditeur CSV/XLS/XLSX</h2>
  <div>Root: <code>{escape(ROOT.as_posix())}</code></div>
  <div>Répertoire courant: <code>{escape(rel_path(target_dir) if target_dir != ROOT else ".")}</code></div>
</div>
<div class="box">
  <ul>{listing}</ul>
</div>
"""
    return Response(page("Éditeur tabulaire", body), mimetype="text/html")


@app.post("/tabular/save")
def tabular_save():
    raw_file = request.args.get("file")
    file_path = safe_target(raw_file)
    if file_path is None or not file_path.exists() or not file_path.is_file():
        return Response(page("Erreur", '<div class="box">Fichier invalide.</div>'), status=400, mimetype="text/html")
    if split_ext(file_path) not in ALLOWED_EXTENSIONS:
        return Response(page("Erreur", '<div class="box">Type de fichier non supporté.</div>'), status=400, mimetype="text/html")
    if not os.access(file_path, os.W_OK):
        return Response(page("Erreur", '<div class="box">Fichier en lecture seule (groupe _ro).</div>'), status=403, mimetype="text/html")

    content = request.form.get("content", "")
    rows = parse_tabular_text(content)
    write_rows(file_path, rows)
    return redirect(f"/tabular/?file={rel_path(file_path)}", code=302)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8090, debug=False)
